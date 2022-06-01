import os
import re
import time
import sys
import getopt
from util.Util import Util
from hashlib import md5
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


# SeqCounter
class SeqCounter:
    def __init__(self):
        self.setting_json = Util.load_setting()
        self.version = self.setting_json['version']
        self.encoding = self.setting_json['seqCounter']['encoding']
        self.seq_path = self.setting_json['seqCounter']['inputOptions']['seqPath']
        self.result_path = self.setting_json['seqCounter']['outputOptions']['resultPath']
        self.time_str = str(time.strftime('%Y%m%d%H%M%S', time.localtime()))
        self.seq_extension_name = self.setting_json['seqCounter']['inputOptions']['seqExtensionName']
        self.result_extension_name = self.setting_json['seqCounter']['outputOptions']['resultExtensionName']
        self.extract_extension_name = self.setting_json['seqCounter']['outputOptions']['extractExtensionName']
        self.save_file = 'result' + self.time_str + self.result_extension_name
        self.seq_type_file_path = self.setting_json['seqCounter']['constraintOptions']['seqTypeList']
        self.virus_info_list = []
        self.check_type_flag = bool(self.setting_json['seqCounter']['constraintOptions']['seqTypeCheck'])
        self.split_symbol_list = self.setting_json['seqCounter']['inputOptions']['symbols']
        self.compare = bool(self.setting_json['seqCounter']['outputOptions']['compare'])
        self.combine_compare = bool(self.setting_json['seqCounter']['outputOptions']['combineCompare'])
        self.extract_seq_flag = bool(self.setting_json['seqCounter']['outputOptions']['extractSeq'])
        self.single_extract = bool(self.setting_json['seqCounter']['outputOptions']['singleExtract'])
        self.extract_extension_name = self.setting_json['seqCounter']['outputOptions']['extractExtensionName']
        self.remove_symbol_list = self.setting_json['seqCounter']['outputOptions']['removeSymbols']
        self.constraint_remove_symbol_list = self.setting_json['seqCounter']['constraintOptions']['removeSymbols']
        self.ignore_empty_seq_flag = bool(self.setting_json['seqCounter']['outputOptions']['ignoreEmptySeq'])
        self.similarity_compare_flag = bool(self.setting_json['seqCounter']['outputOptions']['similarityCompare'])

    # 读取输入文件列表
    def read_path(self):
        if not os.path.isdir(self.seq_path):
            print("错误：指定源文件参数不是目录。")
            return False
        if not os.path.exists(self.seq_path):
            print("错误：指定的源文件目录不存在。")
            return False
        all_file_name_list = os.listdir(self.seq_path)
        if len(all_file_name_list) <= 0:
            print("错误：指定的源文件目录为空。")
            return False
        file_name_list = []
        for file_name in all_file_name_list:
            ex_name_list = re.findall(r'\.[^.\\/:*?"<>|\r\n]+$', file_name)
            if len(ex_name_list) <= 0:
                continue
            ex_name = re.findall(r'\.[^.\\/:*?"<>|\r\n]+$', file_name)[0]
            if ex_name != self.seq_extension_name:
                continue
            file_name_list.append(file_name)
        if len(file_name_list) <= 0:
            print("错误：指定的源文件目录没有任何目标文件。")
            return False
        return file_name_list

    # 读取约束文件列表
    def read_virusinfo(self):
        if not os.path.exists(self.seq_type_file_path):
            print("警告：类型约束文件不存在，将跳过类型判断。")
            return False
        with open(self.seq_type_file_path, 'r', encoding=self.encoding) as file:
            content = file.read()
            if not content or len(content) <= 0:
                print("警告：类型约束文件内容为空，将跳过类型判断。")
                return False
            virus_info_str = Util.remove_char(content, self.constraint_remove_symbol_list)
            if not virus_info_str or len(virus_info_str) <= 0:
                print("警告：类型约束文件内容为空，将跳过类型判断。")
                return False
            pre_virus_info_list = virus_info_str.split(',')
            if not pre_virus_info_list or len(pre_virus_info_list) <= 0:
                print("警告：类型约束文件内容为空，将跳过类型判断。")
                return False
            virus_info_list = []
            index = 0
            for pre_virus_info in pre_virus_info_list:
                index += 1
                if not pre_virus_info or len(pre_virus_info) <= 0:
                    print("警告：第[" + str(index) + ']行参数不合法，已忽略。')
                    continue
                virus_infos = pre_virus_info.split('-')
                if not virus_infos or len(virus_infos) != 2:
                    print("警告：第[" + str(index) + ']行参数不合法，已忽略。')
                    continue
                item = {'virus_name': virus_infos[0]}
                virus_lens = virus_infos[1].split('/')
                if not virus_lens or len(virus_infos) <= 0:
                    print("警告：第[" + str(index) + ']行参数，病毒长度不合法，已忽略。')
                    continue
                item['virus_lens'] = virus_lens
                virus_info_list.append(item)
            return virus_info_list

    # 判断序列类型
    def check_seq_type(self, length):
        if not self.virus_info_list or len(self.virus_info_list) <= 0:
            return '未知'
        for virus_info in self.virus_info_list:
            if str(length) in virus_info['virus_lens']:
                return virus_info['virus_name']
        return '未知'

    # 分割多个序列
    def split_content(self, content):
        symbols = ">"
        if len(self.split_symbol_list) > 0:
            symbols = ""
            for s in self.split_symbol_list:
                symbols += s
        reg = "[" + symbols + "]"
        return re.split(reg, content)

    # 统计序列信息
    def statistics(self):
        file_name_list = self.read_path()
        if not file_name_list:
            print("错误：读取源文件列表失败。")
            return False
        # self.check_type_flag = True
        self.virus_info_list = self.read_virusinfo()
        if not self.virus_info_list or len(self.virus_info_list) <= 0:
            self.check_type_flag = False
        print("开始分析，一共" + str(len(file_name_list)) + "个文件。")
        file_index = 1
        result_list = []
        for file_name in file_name_list:
            full_path = os.path.join(self.seq_path, file_name)
            if not os.path.exists(full_path):
                print("\t警告：源文件[" + full_path + "]不存在。直接跳过。")
                continue
            item = {'file_index': file_index, 'file_name': file_name}
            with open(full_path, 'r', encoding=self.encoding) as file:
                content = file.read()
                if not content or len(content) <= 0:
                    print("\t警告：源文件[" + full_path + "]内容为空。直接跳过。")
                    continue
                seq_str_list = self.split_content(content)
                if len(seq_str_list) <= 0:
                    print("\t警告：源文件[" + full_path + "]中不存在任何序列。直接跳过。")
                    continue
                item['seq_num'] = len(seq_str_list) - 1
            seq_index = 1
            seq_name_reg = '(.*?)\n'
            seq_list = []
            for seq_str in seq_str_list:
                if len(seq_str) <= 0:
                    continue
                seq_name_pattern = re.compile(seq_name_reg, re.S)
                seq_name_list = seq_name_pattern.findall(seq_str)
                if len(seq_name_list) <= 0:
                    # print("警告：第" + str(seq_index) + "个序列解析失败：序列名称不存在。")
                    print("警告：第" + str(file_index) + "个文件[" + file_name + "] 的 第" + str(seq_index) + "个序列解析失败：序列名称不存在。")
                    continue
                seq_item = {}
                seq_name = seq_name_list[0]
                seq_body = seq_str[len(seq_name):]
                seq_body = Util.remove_char(seq_body, self.remove_symbol_list)
                seq_body = seq_body.upper()
                seq_item['seq_index'] = seq_index
                seq_item['seq_name'] = seq_name
                seq_item['seq_length'] = len(seq_body)
                seq_item['seq_body'] = seq_body
                if self.check_type_flag:
                    seq_item['virus_name'] = self.check_seq_type(seq_item['seq_length'])
                seq_list.append(seq_item)
                seq_index = seq_index + 1
            item['seq_list'] = seq_list
            result_list.append(item)
            file_index += 1
        return result_list

    # 合并统计结果为线性列表
    def combine_result(self, result_list):
        if not result_list or len(result_list) <= 0:
            return False
        line_seq_list = []
        for result in result_list:
            seq_list = result['seq_list']
            for seq in seq_list:
                seq['file_index'] = result['file_index']
                seq['file_name'] = result['file_name']
                seq['seq_list'] = result['seq_list']
                line_seq_list.append(seq)
        return line_seq_list

    # 比较序列
    def compare_seq_body(self, line_seq_list):
        if not line_seq_list or len(line_seq_list) <= 0:
            return False
        same_body_map = {}
        for seq in line_seq_list:
            secret = md5()
            secret.update(str(seq['seq_body']).encode())
            finger = secret.hexdigest()
            if finger in same_body_map:
                seq_list = same_body_map[finger]
                seq_list.append(seq)
            else:
                same_body_map[finger] = [seq]
        return same_body_map

    # 打印序列比较结果，输出相同序列信息
    def print_compare(self, result_list):
        compare_info = ""
        have_same_flag = False
        if self.combine_compare:
            compare_info += "\n序列对比结果（模式：跨文件）,以下序列相同：\n"
        else:
            compare_info += "\n序列对比结果（模式：文件内）,以下序列相同：\n"
        if not result_list or len(result_list) <= 0:
            compare_info += "无。\n"
            return compare_info
        if self.combine_compare:
            line_seq_list = self.combine_result(result_list)
            same_body_map = self.compare_seq_body(line_seq_list)
            if not same_body_map:
                compare_info += "无。\n"
                return compare_info
            if not have_same_flag:
                have_same_flag = True
            for key in same_body_map:
                same_seq_list = same_body_map[key]
                if len(same_seq_list) > 1:
                    if self.ignore_empty_seq_flag and len(same_seq_list[0]['seq_body']) <= 0:
                        continue
                    if self.check_type_flag:
                        compare_info += "所在文件：" + same_seq_list[0]['file_name'] + "， 序列类型：" + same_seq_list[0][
                            'virus_name'] + "\n"
                    else:
                        compare_info += "所在文件：" + same_seq_list[0]['file_name'] + "\n"
                    same_seq_str = ""
                    first_falg = True
                    for same_seq in same_seq_list:
                        if first_falg:
                            same_seq_str += same_seq['seq_name']
                            first_falg = False
                            continue
                        same_seq_str += ("," + same_seq['seq_name'])
                    compare_info += "\t相同序列：" + same_seq_str + "\n"
        else:
            for result in result_list:
                line_seq_list = result['seq_list']
                if not line_seq_list or len(line_seq_list) <= 0:
                    continue
                same_body_map = self.compare_seq_body(line_seq_list)
                if not same_body_map:
                    continue
                for key in same_body_map:
                    same_seq_list = same_body_map[key]
                    if len(same_seq_list) > 1:
                        if self.ignore_empty_seq_flag and len(same_seq_list[0]['seq_body']) <= 0:
                            continue
                        if not have_same_flag:
                            have_same_flag = True
                        compare_info += "所在文件：" + result['file_name'] + "\n"
                        same_seq_str = ""
                        first_falg = True
                        for same_seq in same_seq_list:
                            if first_falg:
                                if self.check_type_flag:
                                    compare_info += "序列类型：" + same_seq['virus_name'] + "\n"
                                same_seq_str += same_seq['seq_name']
                                first_falg = False
                                continue
                            same_seq_str += ("," + same_seq['seq_name'])
                        compare_info += "\t相同序列：" + same_seq_str + "\n"
        if not have_same_flag:
            compare_info += "无。" + "\n"
        return compare_info

    # 统计序列相似度并输出表格
    # 相似性=(序列长度-不同的长度)/序列长度
    # 如果A≠B，那么以短的为准，相似性=(短序列长度-不同序列长度)/短序列长度
    def similarity_seq_body(self, result_list):
        if not result_list or len(result_list) <= 0:
            return False
        similar_matrix = []
        for result in result_list:
            result['compared'] = []
            sub_matrix = []
            if not result or not('seq_list' in result) or len(result['seq_list']) <= 0:
                for resultx in result_list:
                    sub_matrix.append(0)
                    result['compared'].append(resultx)
                    if 'compared' in resultx:
                        resultx['compared'].append(result)
                    else:
                        resultx['compared'] = [result]
                similar_matrix.append(sub_matrix)
                continue
            seq_list = result['seq_list']
            seq_body_list = []
            for seq in seq_list:
                seq_body_list.append(seq['seq_body'])
            for resultx in result_list:
                similar_num = 0
                if not('compared' in resultx):
                    resultx['compared'] = []
                # if result in resultx['compared']:
                #     continue
                for seq in resultx['seq_list']:
                    if seq['seq_body'] in seq_body_list:
                        similar_num += 1
                result['compared'].append(resultx)
                resultx['compared'].append(result)
                if len(seq_list) >= len(resultx['seq_list']):
                    if len(seq_list) != 0:
                        sub_matrix.append(round(similar_num / len(seq_list), 5))
                    else:
                        sub_matrix.append(0)
                else:
                    sub_matrix.append(round(similar_num / len(resultx['seq_list']),5))
            similar_matrix.append(sub_matrix)
        # print(similar_matrix)
        return similar_matrix

    # 保存相似度对比结果到xlsx
    def save_similarity_to_xlsx(self, result_list):
        if not result_list or len(result_list) <= 0:
            print("序列分析结果为空：已跳过相似度表格生成。")
            return False
        similar_matrix = self.similarity_seq_body(result_list)
        if not similar_matrix or len(similar_matrix) <= 0:
            print("相似度对比结果为空：已跳过相似度表格生成。")
            return False
        wb = openpyxl.Workbook()
        # ws = wb.active
        st = ""
        if 'Sheet' in wb.sheetnames:
            st = wb['Sheet']
            st.title = "相似性对比"
        else:
            st = wb.create_sheet("相似性对比")
        st.cell(1, 1).alignment = Alignment('center', 'center')
        merge_range = 'A1:' + str(openpyxl.utils.get_column_letter(len(similar_matrix) + 1)) + str(1)
        st.merge_cells(merge_range)
        st.cell(1, 1).value = '相似性对比'
        titleFont = Font(name="黑体", size=12, bold=True)
        st.cell(1, 1).font = titleFont
        headFont = Font(name="Times New Roman", size=11, bold=True)
        contentFont = Font(name="黑体", size=11)
        for x in range(len(result_list)):
            st.cell(2, x + 1).value = Util.get_file_name(result_list[x]['file_name'])
            st.cell(2, x + 1).font = headFont
        for y in range(len(result_list)):
            st.cell(y + 3, 1).value = Util.get_file_name(result_list[y]['file_name'])
            st.cell(y + 3, 1).font = headFont
        # print(similar_matrix)
        for y in range(len(similar_matrix)):
            for x in range(len(similar_matrix)):
                st.cell(x + 3, y + 2).value = round(similar_matrix[x][y] * 100, 2)
                st.cell(x + 3, y + 2).font = contentFont
        similar_excle_name = self.result_path + "Similarity_" + self.time_str + ".xlsx"
        try:
            wb.save(similar_excle_name)
            print("[提示] 序列相似对比结果已保存在[" + similar_excle_name + "]中。")
        except PermissionError as e:
            print("[错误] 序列相似度对比表格保存失败：文件已打开或被其他程序占用。")

    # 提取序列
    def extract_seqs(self, result_list):
        if not result_list or len(result_list) <= 0:
            return False
        seqs_extract_path = self.result_path + "/seqs_extract" + self.time_str + "/"
        if not os.path.exists(seqs_extract_path):
            os.mkdir(seqs_extract_path)
        if not self.single_extract:
            for result in result_list:
                seq_list = result['seq_list']
                if not seq_list or len(seq_list) <= 0:
                    continue
                name = Util.get_file_name(result['file_name']) + self.extract_extension_name
                with open(seqs_extract_path + name, 'w', encoding=self.encoding) as file:
                    for seq in seq_list:
                        # if self.ignore_empty_seq_flag and len(seq['seq_body']) <= 0:
                        #     continue
                        if not seq or not ('seq_name' in seq) or len(seq['seq_name']) <= 0 or not (
                                'seq_body' in seq) or len(seq['seq_body']) <= 0:
                            continue
                        file.write(">" + seq['seq_name'] + "\n" + seq['seq_body'] + "\n")
        else:
            # 单文件（每个序列一个文件）输出，待实现
            # for result in result_list:
            combine_line_seq_list = self.combine_result(result_list)
            if not combine_line_seq_list or len(combine_line_seq_list) <= 0:
                return False
            for seq in combine_line_seq_list:
                if not seq or not ('seq_name' in seq) or len(seq['seq_name']) <= 0 or not ('seq_body' in seq) or len(
                        seq['seq_body']) <= 0:
                    continue
                # if self.ignore_empty_seq_flag and len(seq['seq_body']) <= 0:
                #     continue
                name = Util.get_file_name(seq['file_name']) + "_" + seq['seq_name'].replace("\\", "#").replace("/",
                                                                                                               "#") + self.extract_extension_name
                with open(seqs_extract_path + name, 'w', encoding=self.encoding) as file:
                    file.write(">" + seq['seq_name'] + "\n" + seq['seq_body'] + "\n")
            # return False
        return True

    # 输出序列信息统计结果
    def print_result(self, result_list):
        result_info = ""
        if not result_list or len(result_list) <= 0:
            result_info += "结果为空。\n"
            return result_info
        for result in result_list:
            result_info += "文件" + str(result['file_index']) + "：\n\t文件名：" + result['file_name'] + "\n"
            result_info += "\t序列数：" + str(result['seq_num']) + "\n"
            seq_list = result['seq_list']
            if self.check_type_flag:
                for seq in seq_list:
                    result_info += "\t序号" + str(seq['seq_index']) + "， 序列名称：" + seq['seq_name'] + "， 长度：" + str(
                        seq['seq_length']) + ', 类型：' + seq['virus_name'] + "\n"
            else:
                for seq in seq_list:
                    result_info += "\t序号" + str(seq['seq_index']) + "， 序列名称：" + seq['seq_name'] + "， 长度：" + str(
                        seq['seq_length']) + "\n"
        if self.compare:
            result_info += self.print_compare(result_list)
        return result_info

    # 在控制台打印结果
    def show_result(self, result_list):
        print(self.print_result(result_list))

    # 保存结果到文件
    def save_result(self, result_list):
        if not os.path.exists(self.result_path):
            print("结果保存失败：指定的结果保存目录不存在！")
        result_info = self.print_result(result_list)
        with open(self.result_path + self.save_file, 'w', encoding=self.encoding) as file:
            file.write(result_info)
        if self.extract_seq_flag:
            self.extract_seqs(result_list)

    # 入口函数
    def run(self):
        result_list = self.statistics()
        self.show_result(result_list)
        self.save_result(result_list)
        if self.similarity_compare_flag:
            self.save_similarity_to_xlsx(result_list)
        print()
        print("-" * 50)
        print("统计完毕，结果已保存于[" + self.result_path + self.save_file + "]中。")
        print("请使用任意文本编辑器打开查看。")
        print("-" * 50)

    # 控制台传参
    def get_args(self, argv):
        input_path = self.seq_path
        output_path = self.result_path
        check_path = self.seq_type_file_path
        try:
            opts, args = getopt.getopt(argv, "i:o:c:v:h", ["ifile=", "ofile=", "cfile=", "version", "help"])
        except Exception as e:
            print('SeqCounter.py -i <inputfile> -o <outputfile> -c <checkfile>')
            # print(e)
            sys.exit(2)
        for opt, arg in opts:
            if opt == '-h':
                print('SeqCounter.py -i <inputfile> -o <outputfile> -c <checkfile>')
                sys.exit()
            elif opt in ("-i", "--ifile"):
                input_path = arg
            elif opt in ("-o", "--ofile"):
                output_path = arg
            elif opt in ("-c", "--cfile"):
                check_path = arg
            elif opt in ("-v", "--version"):
                print("SeqCounter version：" + self.version)
                sys.exit(0)
            elif opt in ("-h", "--help"):
                print("SeqCounter Help：\n")
                print("\t[-i] 指定输入文件路径")
                print("\t[-o] 指定输出文件路径")
                print("\t[-c] 指定约束文件路径")
                print("\t[-v] 查看版本号")
                print("\t[-h] 查看帮助信息\n")

                print("本项目以GPL3.0协议开源：https://github.com/nnrj/SeqCounter")
                print("开发团队成员：天河何处")
                sys.exit(0)
        self.seq_path = input_path
        self.result_path = output_path
        self.seq_type_file_path = check_path
        print('输入文件目录：', self.seq_path)
        print('输出文件目录：', self.result_path)
        print('约束文件目录：', self.seq_type_file_path)


# 运行
if __name__ == '__main__':
    seqCounter = SeqCounter()
    seqCounter.get_args(sys.argv[1:])
    seqCounter.run()
